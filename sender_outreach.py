import os
import uuid
import smtplib
from email.message import EmailMessage
from email.utils import formataddr

import pandas as pd
import openpyxl
from dotenv import load_dotenv

from utils import domain_from_email, domain_from_url, normalize_bool, detect_dns_host_from_spf, safe_company_from_domain
from templates import render_template
from db import record_message

load_dotenv()

SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASS = os.getenv("SMTP_PASS", "")

SENDER_NAME = os.getenv("SENDER_NAME", "Technology Transition Paradigm")
SENDER_EMAIL = os.getenv("SENDER_EMAIL", SMTP_USER)
REPLY_TO_EMAIL = os.getenv("REPLY_TO_EMAIL", SENDER_EMAIL)

UNSUBSCRIBE_EMAIL = os.getenv("UNSUBSCRIBE_EMAIL", "")
PHONE = os.getenv("CALL_TO_ACTION_PHONE", "800-889-8072")

PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "").rstrip("/")
TRACKER_SECRET = os.getenv("TRACKER_SECRET", "")

RESULTS_XLSX = os.getenv("RESULTS_XLSX", "ttpResults.xlsx")
APOLLO_CSV = os.getenv("APOLLO_CSV", "apollo.csv")

MAX_EMAILS_PER_RUN = int(os.getenv("MAX_EMAILS_PER_RUN", "50"))
DRY_RUN = os.getenv("DRY_RUN", "false").strip().lower() == "true"

def scenario_from_bools(spf_ok: bool, dmarc_ok: bool, dkim_ok: bool, dmarc_weak: bool) -> int:
    spf_missing = not spf_ok
    dkim_missing = not dkim_ok
    dmarc_missing = (not dmarc_ok) and (not dmarc_weak)

    if spf_missing and dkim_missing and (dmarc_missing or dmarc_weak or not dmarc_ok):
        return 5
    if spf_missing and dmarc_missing and dkim_ok:
        return 3
    if spf_missing and dkim_ok and dmarc_ok:
        return 2
    if dkim_missing and spf_ok and dmarc_ok:
        return 4
    if dmarc_weak and spf_ok and dkim_ok:
        return 1
    return 0

def build_col_map(ws):
    headers = [c.value for c in ws[1]]
    return {h: i+1 for i, h in enumerate(headers) if h}

def set_cell(ws, row, col_map, header, value):
    col = col_map.get(header)
    if col:
        ws.cell(row=row, column=col).value = value

def send_email(to_email: str, subject: str, plain: str, html: str, message_id: str, sender_domain: str):
    k = f"&k={TRACKER_SECRET}" if TRACKER_SECRET else ""
    pixel_url = f"{PUBLIC_BASE_URL}/pixel/{message_id}.png?k=1{k}" if TRACKER_SECRET else f"{PUBLIC_BASE_URL}/pixel/{message_id}.png"
    html_tracked = html.replace("</body>", f'<img src="{pixel_url}" width="1" height="1" alt="" /></body>')

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = formataddr((SENDER_NAME, SENDER_EMAIL))
    msg["To"] = to_email
    msg["Reply-To"] = REPLY_TO_EMAIL

    if UNSUBSCRIBE_EMAIL:
        msg["List-Unsubscribe"] = f"<mailto:{UNSUBSCRIBE_EMAIL}?subject=unsubscribe>"

    msg["Message-ID"] = f"<{message_id}@{sender_domain}>"
    msg["X-Auto-Generated"] = "true"

    msg.set_content(plain)
    msg.add_alternative(html_tracked, subtype="html")

    if DRY_RUN:
        print(f"[DRY_RUN] Would send to {to_email}: {subject}")
        return

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)

def main():
    if not (SMTP_HOST and SMTP_USER and SMTP_PASS and SENDER_EMAIL):
        raise RuntimeError("Missing SMTP env vars. Check .env")

    if not PUBLIC_BASE_URL:
        raise RuntimeError("Set PUBLIC_BASE_URL in .env (your tracker domain).")

    contacts = pd.read_csv(APOLLO_CSV)

    by_domain = {}
    for _, r in contacts.iterrows():
        email = r.get("Email", "")
        website = r.get("Website", "")
        dom = domain_from_email(str(email)) or domain_from_url(str(website))
        if dom and dom not in by_domain:
            by_domain[dom] = r

    wb = openpyxl.load_workbook(RESULTS_XLSX)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
    col_map = build_col_map(ws)

    sender_domain = domain_from_email(SENDER_EMAIL) or "local"
    sent = 0

    for row in range(2, ws.max_row + 1):
        if sent >= MAX_EMAILS_PER_RUN:
            break

        website = ws.cell(row=row, column=col_map["Website"]).value
        domain = domain_from_url(str(website)) if website else None
        if not domain:
            continue

        if normalize_bool(ws.cell(row=row, column=col_map["Email Sent"]).value):
            continue

        spf_ok = normalize_bool(ws.cell(row=row, column=col_map["SPF"]).value)
        dmarc_ok = normalize_bool(ws.cell(row=row, column=col_map["DMARC"]).value)
        dkim_ok = normalize_bool(ws.cell(row=row, column=col_map["DKIM"]).value)

        # Your scanner logs DMARC weak as ISSUE, but the sheet stores DMARC boolean only.
        # For now, we treat DMARC weak as False unless you add a 'DMARC Status' column.
        dmarc_weak = False

        spf_record = ws.cell(row=row, column=col_map["SPF Record Description"]).value
        dns_host = detect_dns_host_from_spf(str(spf_record) if spf_record else None)

        scenario = scenario_from_bools(spf_ok, dmarc_ok, dkim_ok, dmarc_weak)

        contact = by_domain.get(domain)
        if contact is None:
            parts = domain.split(".")
            for i in range(1, min(3, len(parts))):
                cand = ".".join(parts[i:])
                if cand in by_domain:
                    contact = by_domain[cand]
                    break
        if contact is None:
            continue

        to_email = str(contact.get("Email", "")).strip()
        if "@" not in to_email:
            continue

        first = str(contact.get("First Name", "")).strip()
        last = str(contact.get("Last Name", "")).strip()
        title = str(contact.get("Title", "")).strip()
        company = str(contact.get("Company Name", "") or safe_company_from_domain(domain)).strip()

        subject, plain, html = render_template(scenario, first, domain, dns_host, PHONE, company)

        message_id = str(uuid.uuid4())
        send_email(to_email, subject, plain, html, message_id, sender_domain)
        record_message(message_id, to_email, domain, scenario)

        set_cell(ws, row, col_map, "First Name", first)
        set_cell(ws, row, col_map, "Last Name", last)
        set_cell(ws, row, col_map, "Title", title)
        set_cell(ws, row, col_map, "Company", company)
        set_cell(ws, row, col_map, "Company Name for Emails", company)
        set_cell(ws, row, col_map, "Email", to_email)

        set_cell(ws, row, col_map, "Email Sent", True)
        set_cell(ws, row, col_map, "Email Open", False)
        set_cell(ws, row, col_map, "Email Bounced", False)
        set_cell(ws, row, col_map, "Replied", False)

        # Optional custom columns (if you add them to Excel)
        if "Message ID" in col_map:
            set_cell(ws, row, col_map, "Message ID", message_id)
        if "Scenario" in col_map:
            set_cell(ws, row, col_map, "Scenario", scenario)

        sent += 1
        print(f"Sent {sent}: {domain} -> {to_email} (scenario {scenario})")

    wb.save(RESULTS_XLSX)
    print("Done.")

if __name__ == "__main__":
    main()
