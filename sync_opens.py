import os
import sqlite3

import openpyxl
from dotenv import load_dotenv

from utils import normalize_bool, domain_from_url

load_dotenv()

DB_PATH = os.getenv("TRACK_DB_PATH", "email_tracking.db")
RESULTS_XLSX = os.getenv("RESULTS_XLSX", "ttpResults.xlsx")

def build_col_map(ws):
    headers = [c.value for c in ws[1]]
    return {h: i+1 for i, h in enumerate(headers) if h}

def set_cell(ws, row, col_map, header, value):
    col = col_map.get(header)
    if col:
        ws.cell(row=row, column=col).value = value

def main():
    with sqlite3.connect(DB_PATH) as conn:
        opened = {r[0] for r in conn.execute(
            "SELECT DISTINCT message_id FROM events WHERE event_type='open'"
        ).fetchall()}

        msg_map = {r[0]: (r[1], r[2]) for r in conn.execute(
            "SELECT message_id, email, domain FROM messages"
        ).fetchall()}

    if not opened:
        print("No opens recorded yet.")
        return

    wb = openpyxl.load_workbook(RESULTS_XLSX)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
    col_map = build_col_map(ws)

    have_mid = "Message ID" in col_map
    updated = 0

    if have_mid:
        for row in range(2, ws.max_row + 1):
            mid = ws.cell(row=row, column=col_map["Message ID"]).value
            if isinstance(mid, str) and mid in opened:
                if not normalize_bool(ws.cell(row=row, column=col_map["Email Open"]).value):
                    set_cell(ws, row, col_map, "Email Open", True)
                    updated += 1
    else:
        opened_pairs = {(msg_map[mid][0], msg_map[mid][1]) for mid in opened if mid in msg_map}
        for row in range(2, ws.max_row + 1):
            email = ws.cell(row=row, column=col_map["Email"]).value
            website = ws.cell(row=row, column=col_map["Website"]).value
            dom = domain_from_url(str(website)) if website else None
            if isinstance(email, str) and dom and (email, dom) in opened_pairs:
                if not normalize_bool(ws.cell(row=row, column=col_map["Email Open"]).value):
                    set_cell(ws, row, col_map, "Email Open", True)
                    updated += 1

    wb.save(RESULTS_XLSX)
    print(f"Updated Email Open for {updated} rows.")

if __name__ == "__main__":
    main()
